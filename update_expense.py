import openpyxl
import sys
from datetime import date
from openpyxl.styles import Font
fileName = 'C:\\Karthik\\Expenses\\Expenses-2020.xlsx'
wb=openpyxl.load_workbook(fileName)

def insertExpenseItem(thisMonth,inputArray):
    sheet=wb[thisMonth]
    currentRow = str(sheet.max_row + 1)
    sheet['A'+currentRow] = date.today()
    sheet['B'+currentRow] = inputArray[0]
    sheet['C'+currentRow] = float(inputArray[1])
    wb.save(fileName)

def update(inputValue):
    inputArray = inputValue.split()  
    thisMonth = date.today().strftime("%B")
    if(thisMonth in wb.sheetnames):
        insertExpenseItem(thisMonth,inputArray)
    else:
        wb.create_sheet(thisMonth)
        sheet=wb[thisMonth]
        sheet["A1"]="Date"
        sheet["B1"] = "Expense Item"
        sheet["C1"] = "Amount"
        fontObject = Font(bold=True)
        sheet['A1'].font = fontObject
        sheet['B1'].font = fontObject
        sheet['C1'].font = fontObject
        wb.save(fileName)
        insertExpenseItem(thisMonth,inputArray)

if __name__ == '__main__':
    print(sys.argv[1])
    update(sys.argv[1])